import os
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook
from numpy import asarray
from collections import defaultdict
from PIL import Image
from numpy import asarray
import numpy as np
from tqdm import tqdm
import gdown
from pydub import AudioSegment
import random
import json
import librosa

from utils.nlp import clean_text
from utils.process import process_audio
from utils.split import custom_train_test_split


def load_data(file_name, folder_name=None):
    parent_path = Path(os.getcwd()).parent
    if folder_name:
        # path = parent_path / folder_name / file_name
        path = Path("/".join([os.getcwd(), folder_name, file_name]))
    else:
        # path = parent_path / file_name
        path = Path("/".join([os.getcwd(), file_name]))
    
    file_type = file_name.split(".")[1]
    if file_type=="csv":
        output = pd.read_csv(path)
    elif file_type== "xlsx":
        output = pd.read_excel(path)
    else:
        raise f"Failed to load data - invalid file type {file_type}"
    
    print(output.info(), "\n")
    print(output.describe(), "\n")

    return output


def check_file_downloaded(file_name, download_dir):
    file_path = download_dir / file_name
    if os.path.exists(file_path):
        print(f"File {file_name} already exists in {download_dir}.")
        return True
    else:
        print(f"File {file_name} doesn't exist in {download_dir}.")
        return False


def extract_zip_file(file_path, download_path, file_name):
    if os.path.exists(download_path + "\\" + file_name.split(".")[0]):
        print(f"{file_name} already extracted in {download_path}.")
    else:
        with ZipFile(file_path, 'r') as zip_file:
            zip_file.extractall(path=download_path)
        print(f"{file_name} extracted in {download_path}.")


def load_images(download_path, as_array=False):
    files_dict = defaultdict(dict)
    for f1 in os.listdir(download_path):
        if f1 == "images":
            print(f"Loading files in {download_path}\\{f1}")
            for f2 in os.listdir(download_path + f"\\{f1}"): # training or testing
                if files_dict.get(f2, "") == "":
                    files_dict[f2] = defaultdict(dict)

                print(f"Loading files in {download_path}\\{f1}\\{f2}")
                for f3 in os.listdir(download_path + f"\\{f1}\\{f2}"): # flip or notflip
                    if files_dict.get(f3, "") == "":
                        files_dict[f3] = defaultdict(dict)
                    
                    print(f"Loading files in {download_path}\\{f1}\\{f2}\\{f3}")
                    for f4 in tqdm(os.listdir(download_path + f"\\{f1}\\{f2}\\{f3}")):
                        if files_dict.get(f4, "") == "":
                            files_dict[f4] = defaultdict(dict)
                        
                        # Load each image file and convert it into a 3d (RGB) array.
                        jpg_file_path = download_path + f"\\{f1}\\{f2}\\{f3}\\{f4}"
                        image = Image.open(jpg_file_path)
                        if as_array:
                            image_array = asarray(image)
                            files_dict[f2][f3][f4] = image_array
                        else:
                            files_dict[f2][f3][f4] = image

    return files_dict


def get_image_shape(array_dict):
    image_shape = None
    for k, v in array_dict.items():
        for k2, v2 in v.items():
            for k3, v3 in v2.items():
                while image_shape == None:
                    image_array = asarray(v3)
                    image_shape = image_array.shape
                    print(f"Image shape: {image_shape}")
    return image_shape


def load_dataframes(file_path, full_sheet_names):
    data_dfs = {}
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    for full_sheet_name, sheet_name in zip(full_sheet_names, sheet_names):
        sheet= wb[sheet_name]
        data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        data_dfs[full_sheet_name] = pd.DataFrame(data[1:-1], columns=data[0])

    print(f'{len(data_dfs)} DataFrames loaded with the following sheet names:\n')
    for sheet_name in full_sheet_names:
        print(sheet_name)

    return data_dfs


def download_from_gdrive(file_id, file_name, download_dir, root_dir):
  if not check_file_downloaded(file_name, download_dir):
    os.chdir(download_dir)
    gdown.download(id=file_id, output=file_name)
    os.chdir(root_dir)


def get_variables_for_voice_cloning(source_audio_subpath, target_audio_subpath,
                                    root_dir, timit_dir, df):
    # Get source details
    source_speaker_id = source_audio_subpath.split('/')[2]
    source_audio_file = source_audio_subpath.split('/')[3]
    source_file_id = source_audio_file.split('.')[0]
    source_audio_path = timit_dir / 'data' / source_audio_subpath
    print(f'Source path: {source_audio_path}')

    # Get source text
    source_text_subpath = df[(df['speaker_id']==source_speaker_id) &
            (df['filename']==source_file_id+'.TXT')]['path_from_data_dir'].iloc[0]
    source_text_file = source_text_subpath.split('/')[3]
    source_text_path = timit_dir / 'data' / source_text_subpath
    with open(source_text_path) as txt:
        raw_source_text = txt.read().split()[2:]
    source_text = clean_text(raw_source_text, remove_whitespace=True, remove_punctuation=True, lower=True)

    # Get target details
    target_speaker_id = target_audio_subpath.split('/')[2]
    target_audio_file = target_audio_subpath.split('/')[3]
    target_file_id = target_audio_file.split('.')[0]
    target_audio_path = timit_dir / 'data' / target_audio_subpath
    print(f'Target path: {target_audio_path}')

    # Get output details
    output_folder = root_dir / 'output' / f'{source_speaker_id}-{source_file_id}_to_{target_speaker_id}-{target_file_id}'
    output_filename = f'{source_speaker_id}-{source_file_id}_to_{target_speaker_id}-{target_file_id}.wav'

    return source_speaker_id, source_audio_file, source_file_id, source_audio_path,\
        target_speaker_id, target_audio_file, target_file_id, target_audio_path,\
        source_text_file, source_text_path, source_text, output_folder, output_filename


def get_concat_audio(target_audio_path, target_concat_dir, target_speaker_id,
                    add_silence=0, n_duplicate_concat=1):
    target_parent_dir = os.path.split(target_audio_path)[0]
    target_output_path = target_concat_dir / f'{target_speaker_id}_concat.wav'
    
    if os.path.exists(target_output_path):
        print(f'{target_output_path} already exists.')
    
    else:
        print(f'Creating a concatenated audio file for target speaker {target_speaker_id}.')
        concat_audio = AudioSegment.empty()
        file_list = [file for file in os.listdir(target_parent_dir) if 'WAV.wav' in file]

        for i in range(n_duplicate_concat):
            random.shuffle(file_list)
            for file in file_list:
                wav_path = target_parent_dir / Path(file)
                audio = AudioSegment.from_wav(wav_path) + AudioSegment.silent(add_silence) # default: 1000ms=1 second
                concat_audio += audio

        concat_audio.export(target_output_path, 'wav')
    
    return target_output_path


def get_stratified_sample(df_path, df, groupby, label_col, num_sample_per_lebel):
    if os.path.exists(df_path):
        stratified_sample_df = pd.read_csv(df_path)
        print('stratified_sample_df loaded from CSV.')

    else:
        stratified_sample = df.groupby(groupby)[label_col].apply(
            lambda x: x.sample(num_sample_per_lebel)).reset_index(drop=True).values
        stratified_sample_df = df[ df[label_col].apply(lambda x: x in stratified_sample) ]
        stratified_sample_df.to_csv(df_path, index=False)
        print('stratified_sample_df saved as CSV.')

    return stratified_sample_df


def get_X_and_y(train_X_path, test_X_path, train_y_path, test_y_path, audio_path, audio_dir,
                label_col='label', n_mfcc=40, train_size=0.8):
    if os.path.exists(train_X_path) and\
        os.path.exists(test_X_path) and\
        os.path.exists(train_y_path) and\
        os.path.exists(test_y_path):

        # Load CSV files to data frames.
        train_X = pd.read_csv(train_X_path)
        test_X = pd.read_csv(test_X_path)
        train_y = pd.read_csv(train_y_path)[label_col]
        test_y = pd.read_csv(test_y_path)[label_col]
        print('CSV files already exist. Skipping processing audio data.')
        print('CSV files loaded as data frames.')

    else:
        print('CSV files do not exist. Processing audio data.')
        # Process audio files to get train_df.
        _, _, train_df = process_audio(audio_path, audio_dir, n_mfcc=n_mfcc)

        # Split train_df into train and test sets.
        _, train_X, train_y, _, test_X, test_y\
            = custom_train_test_split(train_df, train_size=train_size)

        # Save data frames to CSV files.
        train_X.to_csv(train_X_path, index=False)
        test_X.to_csv(test_X_path, index=False)
        train_y.to_csv(train_y_path, index=False)
        test_y.to_csv(test_y_path, index=False)
        print('Data frames saved as CSV files.')

    return train_X, test_X, train_y, test_y


def get_label_dict(label_dict_path, unique_labels):
    if os.path.exists(label_dict_path):
        print(f'Existing label dict loaded from {label_dict_path}.')
        label_dict = json.load(open(label_dict_path, 'r'))

    else:
        label_dict = {label: i for i, label in enumerate(unique_labels)}
        json.dump(label_dict, open(label_dict_path, 'w'))
        print(f'label dict created and saved in {label_dict_path}.')
        
    reverse_label_dict = {v: k for k, v in label_dict.items()}

    return label_dict, reverse_label_dict


def get_cloned_audio_paths(main_output_dir):
    cloned_audio_paths = []
    for folder in os.listdir(main_output_dir):
        if '_to_' in folder:
            for file in os.listdir(main_output_dir / folder):
                if '_to_' in file:
                    file_path = main_output_dir / folder / file
                    cloned_audio_paths.append(file_path)
    
    return cloned_audio_paths


def get_X_y_from_cloned_audio(main_output_dir, n_mfcc, label_col):
    X_list = []
    y_list = []
    wav_path_list = []
    for folder in os.listdir(main_output_dir):
        for file in os.listdir(main_output_dir / folder):
            if '_to_' in file:
                # if 'vc_' in file: # This is to only include files generated by speech_generator from voice_cloning library
                wav_path = main_output_dir / folder / file
                wav_path_list.append(wav_path)
                
                label = file.split('_to_')[1].split('-')[0]
                y_list.append(label)

                audio_array, sample_rate = librosa.load(wav_path, sr=None)
                # print(f'Shape of audio array: {audio_array.shape}')
                mfccs_features = librosa.feature.mfcc(y=audio_array, sr=sample_rate, n_mfcc=n_mfcc)
                # print(f'Shape of mfccs features: {mfccs_features.shape}')
                normalized_mfccs_features = np.mean(mfccs_features.T, axis=0)
                # print(f'Shape of scaled mfccs features: {normalized_mfccs_features.shape}')
                X_list.append(normalized_mfccs_features)

    df = pd.DataFrame(X_list)
    df[label_col] = y_list

    X = df.drop([label_col], axis=1)
    y = df[[label_col]]

    return X, y, wav_path_list


def get_pred_df(model, X, y, label_col, speaker_id_dict, reverse_speaker_id_dict):
    pred_code = np.argmax(model.predict(X), axis=1)
    pred_prob = np.max(model.predict(X), axis=1)
    label_encoded = [speaker_id_dict[label] for label in y['label']]
    pred_label = [reverse_speaker_id_dict[code] for code in pred_code]

    pred_df = y.copy()
    pred_df['code'] = label_encoded
    pred_df['pred_label'] = pred_label
    pred_df['pred_code'] = pred_code
    pred_df['pred_prob'] = pred_prob
    pred_df['matched'] = (pred_df[label_col] == pred_df['pred_label'])

    return pred_df


def get_wav_df(wav_path_list):
    wav_dict_list = []
    for wav_path in wav_path_list:
        substr = str(wav_path).split('output')[1].split('\\')[2]
        if '_M' in substr.split('_to_')[0]:
            model_name = substr.split('_to_')[0].split('_M')[0]
        else:
            model_name = substr.split('_to_')[0].split('_F')[0]
        subsubstr = substr.split(model_name+'_')[1].replace('.wav', '')
        source, target = subsubstr.split('_to_')
        source_speaker, source_script = source.split('-')
        target_speaker, target_script = target.split('-')
        wav_dict = {
            'model_name': model_name,
            'source_speaker': source_speaker,
            'source_script': source_script,
            'target_speaker': target_speaker,
            'target_script': target_script,
            'wav_path': wav_path
        }
        wav_dict_list.append(wav_dict)
    wav_df = pd.DataFrame(wav_dict_list)

    return wav_df


def get_result_df(model, normalized_X, y, label_col, speaker_id_dict, reverse_speaker_id_dict, wav_path_list):
    # Get pred_df
    pred_df = get_pred_df(model, normalized_X, y, label_col, speaker_id_dict, reverse_speaker_id_dict)

    # Get wav_df
    wav_df = get_wav_df(wav_path_list)

    # Concatenate pred_dif and wav_df to get result_df
    result_df = pd.concat([pred_df, wav_df], axis=1)

    return result_df